from celery import shared_task
from django.conf import settings
import logging
import os

logger = logging.getLogger(__name__)

try:
    from adapters.realtime.chatbot_logger import ids_log, IDS
except ImportError:
    def ids_log(*a, **kw): pass
    class IDS:
        TASK='TASK'; ERROR='ERROR'


@shared_task(queue='reportes', bind=True, max_retries=2)
def generar_informe_pdf(self, certificacion_id):


    from apps.certificaciones.models import Certificacion

    ids_log(IDS.TASK, msg='generar_pdf_start', cert_id=str(certificacion_id))
    try:
        cert = Certificacion.objects.select_related(
            'expediente__cliente',
            'expediente__tipo_auditoria',
            'expediente__auditor_lider',
        ).get(id=certificacion_id)

        html_content = _render_certificado_html(cert)
        pdf_bytes    = _html_to_pdf(html_content)

        nombre_archivo = f'{cert.numero.replace("/", "-")}.pdf'
        ruta_relativa  = f'certificados/{nombre_archivo}'
        ruta_absoluta  = os.path.join(settings.MEDIA_ROOT, 'certificados', nombre_archivo)

        os.makedirs(os.path.dirname(ruta_absoluta), exist_ok=True)
        with open(ruta_absoluta, 'wb') as f:
            f.write(pdf_bytes)

        cert.certificado_pdf = ruta_relativa
        cert.estado = 'VIGENTE'
        cert.save(update_fields=['certificado_pdf', 'estado'])

        _notificar_pdf_listo(cert)
        ids_log(IDS.TASK, msg='generar_pdf_ok', path=ruta_relativa)
        logger.info('PDF generado: %s', ruta_absoluta)
        return str(ruta_relativa)

    except (RuntimeError, ImportError) as exc:

        logger.error('Error permanente generando PDF %s: %s', certificacion_id, exc)
        raise

    except Exception as exc:
        logger.error('Error transitorio generando PDF %s: %s', certificacion_id, exc)
        raise self.retry(exc=exc, countdown=60)


def _render_certificado_html(cert):

    from django.template.loader import render_to_string
    exp = cert.expediente

    hallazgos = exp.hallazgos.order_by('nivel_criticidad').values(
        'titulo', 'nivel_criticidad', 'estado', 'descripcion'
    )
    checklist = exp.checklist.order_by(
        'item__fase__orden', 'item__orden'
    ).select_related('item__fase')

    equipo = exp.equipo.filter(activo=True).select_related('usuario')
    docs   = exp.documentos.all().select_related()

    context = {
        'cert':       cert,
        'expediente': exp,
        'cliente':    exp.cliente,
        'tipo':       exp.tipo_auditoria,
        'hallazgos':  list(hallazgos),
        'checklist':  list(checklist),
        'equipo':     list(equipo),
        'documentos': list(docs),
        'base_url':   getattr(settings, 'FRONTEND_URL', 'http://localhost:3000'),
        'media_url':  settings.MEDIA_URL,
    }
    return render_to_string('pdf/certificado.html', context)


def _html_to_pdf(html_content):

    try:
        from weasyprint import HTML
        from weasyprint.text.fonts import FontConfiguration
        font_config = FontConfiguration()
        return HTML(string=html_content).write_pdf(font_config=font_config)
    except ImportError:
        raise RuntimeError(
            'WeasyPrint no está instalado. '
            'Ejecuta: pip install weasyprint --break-system-packages'
        )


def _notificar_pdf_listo(cert):

    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        lider = cert.expediente.auditor_lider

        if not lider or not getattr(lider, 'id', None):
            return
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f'notificaciones_{lider.id}',
            {
                'type':    'notificacion',
                'tipo':    'INFO',
                'titulo':  f'PDF listo — {cert.numero}',
                'mensaje': f'El certificado PDF de {cert.expediente.cliente.razon_social} está disponible.',
            },
        )
    except Exception as e:
        logger.warning('No se pudo notificar PDF listo: %s', e)


@shared_task(queue='reportes')
def generar_reporte_excel(tipo, filtros=None):


    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone

    ids_log(IDS.TASK, msg='generar_excel_start', tipo=tipo)
    filtros = filtros or {}


    user_id = filtros.pop('_user_id', None)

    wb = openpyxl.Workbook()
    ws = wb.active

    NAVY  = 'FF0F2447'
    WHITE = 'FFFFFFFF'
    LGRAY = 'FFF1F5F9'

    header_font  = Font(name='Calibri', bold=True, color=WHITE, size=11)
    header_fill  = PatternFill('solid', fgColor=NAVY)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin', color='FFCCCCCC'),
        right=Side(style='thin', color='FFCCCCCC'),
        bottom=Side(style='thin', color='FFCCCCCC'),
    )

    dispatch = {
        'expedientes':    (_excel_expedientes,    'Expedientes'),
        'hallazgos':      (_excel_hallazgos,      'Hallazgos'),
        'certificaciones':(_excel_certificaciones,'Certificaciones'),
        'clientes':       (_excel_clientes,       'Clientes'),
    }

    if tipo not in dispatch:
        raise ValueError(f'Tipo de reporte desconocido: {tipo}')

    fn, titulo = dispatch[tipo]
    fn(ws, filtros, header_font, header_fill, header_align, thin_border, LGRAY)
    ws.title = titulo

    ts     = timezone.now().strftime('%Y%m%d_%H%M%S')
    nombre = f'reporte_{tipo}_{ts}.xlsx'
    ruta   = os.path.join(settings.MEDIA_ROOT, 'reportes', nombre)
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    wb.save(ruta)
    ids_log(IDS.TASK, msg='generar_excel_ok', tipo=tipo, nombre=nombre)
    logger.info('Reporte Excel generado: %s', ruta)


    if user_id:
        try:
            from asgiref.sync import async_to_sync
            from channels.layers import get_channel_layer
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'notificaciones_{user_id}',
                {
                    'type':    'notificacion',
                    'tipo':    'INFO',
                    'titulo':  f'Reporte {titulo} listo',
                    'mensaje': f'Tu reporte Excel está disponible para descargar.',
                    'url':     f'/media/reportes/{nombre}',
                },
            )
        except Exception as e:
            logger.warning('No se pudo notificar reporte listo a user %s: %s', user_id, e)

    return f'reportes/{nombre}'


def _set_headers(ws, headers, header_font, header_fill, header_align):
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 30


def _excel_expedientes(ws, filtros, hf, hfill, ha, border, lgray):
    from apps.expedientes.models import Expediente
    _set_headers(ws, ['Número', 'Cliente', 'NIT', 'Tipo Auditoría', 'Estado',
                      'Auditor Líder', 'Fecha Apertura', '% Avance', 'Cierre Estimado'], hf, hfill, ha)
    qs = Expediente.objects.select_related('cliente', 'tipo_auditoria', 'auditor_lider')
    if filtros.get('estado'):
        qs = qs.filter(estado=filtros['estado'])
    for exp in qs:
        ws.append([
            exp.numero_expediente,
            exp.cliente.razon_social,
            exp.cliente.nit,
            exp.tipo_auditoria.nombre,
            exp.get_estado_display(),
            exp.auditor_lider.nombre_completo if exp.auditor_lider else '',
            exp.fecha_apertura.strftime('%d/%m/%Y') if exp.fecha_apertura else '',
            float(exp.porcentaje_avance),
            exp.fecha_estimada_cierre.strftime('%d/%m/%Y') if exp.fecha_estimada_cierre else '',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18


def _excel_hallazgos(ws, filtros, hf, hfill, ha, border, lgray):
    from apps.ejecucion.models import Hallazgo
    _set_headers(ws, ['ID', 'Expediente', 'Cliente', 'Tipo', 'Criticidad',
                      'Título', 'Estado', 'Fecha Registro'], hf, hfill, ha)
    qs = Hallazgo.objects.select_related('expediente__cliente').order_by('nivel_criticidad', '-fecha_creacion')
    if filtros.get('criticidad'):
        qs = qs.filter(nivel_criticidad=filtros['criticidad'])
    for hallazgo in qs:
        ws.append([
            str(hallazgo.id)[:8],
            hallazgo.expediente.numero_expediente,
            hallazgo.expediente.cliente.razon_social,
            hallazgo.get_tipo_display(),
            hallazgo.get_nivel_criticidad_display(),
            hallazgo.titulo,
            hallazgo.get_estado_display(),
            hallazgo.fecha_creacion.strftime('%d/%m/%Y') if hallazgo.fecha_creacion else '',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20


def _excel_certificaciones(ws, filtros, hf, hfill, ha, border, lgray):
    from apps.certificaciones.models import Certificacion
    _set_headers(ws, ['Número', 'Cliente', 'Tipo Auditoría', 'Estado',
                      'Fecha Emisión', 'Fecha Vencimiento', 'Código Verificación'], hf, hfill, ha)
    qs = Certificacion.objects.select_related('expediente__cliente', 'expediente__tipo_auditoria', 'expediente__auditor_lider')
    if filtros.get('estado'):
        qs = qs.filter(estado=filtros['estado'])
    for cert in qs:
        ws.append([
            cert.numero,
            cert.expediente.cliente.razon_social,
            cert.expediente.tipo_auditoria.nombre,
            cert.get_estado_display(),
            cert.fecha_emision.strftime('%d/%m/%Y') if cert.fecha_emision else '',
            cert.fecha_vencimiento.strftime('%d/%m/%Y') if cert.fecha_vencimiento else '',
            cert.codigo_verificacion,
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22


def _excel_clientes(ws, filtros, hf, hfill, ha, border, lgray):
    from apps.clientes.models import Cliente
    _set_headers(ws, ['Razón Social', 'NIT', 'Sector', 'Ciudad', 'Estado',
                      'Email', 'Teléfono', 'Representante Legal'], hf, hfill, ha)
    qs = Cliente.objects.all()
    if filtros.get('estado'):
        qs = qs.filter(estado=filtros['estado'])
    for c in qs:
        ws.append([
            c.razon_social, c.nit, c.get_sector_display(),
            c.ciudad, c.get_estado_display(),
            c.email, c.telefono, c.rep_legal_nombre,
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
